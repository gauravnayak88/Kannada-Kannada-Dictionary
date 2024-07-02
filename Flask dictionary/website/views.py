from flask import Blueprint, render_template, request, redirect
import openpyxl
workbook = openpyxl.load_workbook('website/static/dictionary.xlsx')
workbook_dictionary=workbook['Sheet1']
univ_dict = {}
for row in workbook_dictionary.iter_rows(min_row=4, values_only=True):
    title, meaning= row[0], row[1]
    univ_dict[title]=meaning

views= Blueprint('views', __name__)

@views.route('/', methods=['GET', 'POST'])
def home():
    return render_template('index.html')

@views.route('/allwords', methods=['GET', 'POST'])
def allwords():
    return render_template('index.html', words=univ_dict, all=True)

    
@views.route('/searchword', methods=['GET', 'POST'])
def searchword():
    if request.method=='POST':
        sub=request.form.get('title')
        print(sub)
        if (sub==""):
            return redirect('/')
        selected_words={}
        selected_meanings={}
        for row in workbook_dictionary.iter_rows(min_row=2, values_only=True):
            title, meaning= row[0], row[1]
            if sub in title:
                selected_words[title]=meaning
                print(title, meaning)
            if sub in meaning:
                selected_meanings[title]=meaning
                print(title, meaning)
        return render_template('index.html', words=selected_words, meanings=selected_meanings, search=True)
    return render_template('index.html')

@views.route('/about', methods=['GET', 'POST'])
def about():
    return render_template('about.html')